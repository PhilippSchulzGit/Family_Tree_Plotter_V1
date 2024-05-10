# -*- coding: utf-8 -*-
"""
Created on Sat Oct 28 19:49:25 2023

Code for extracting data from an excel file for plotting family trees

@author: Philipp Schulz
"""
# imports
from openpyxl import load_workbook
import os

# global variables
letters = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q",
           "R","S","T","U","V","W","X","Y","Z"]

def extract_values_from_sheet(sheet):
    """
    Extracts the content on a given sheet and returns it as a list.
    @param sheet: Worksheet from which the data shall be extracted.
    @return: Content of the worksheet in form of a list of list of strings.
    """
    # initialize return value
    worksheet_data = []
    # loop over first row to determine number of fields
    fields = []
    for i in range(0,sheet.max_column,1):
        # get current cell
        cell = sheet[str(letters[i])+"1"].value
        # check if cell is empty
        if(len(str(cell))>0 and str(cell) != "None"):
            # save new field
            fields.append(str(cell))
    # add fields to return list
    worksheet_data = [[fields[i]] for i in range(0, len(fields), 1)]
    # read people data from first sheet
    for i in range(2, sheet.max_row+1, 1):
        for j in range(0, len(fields), 1):
            # assign current row content to return list
            worksheet_data[j].append(str(sheet[str(letters[j])+str(i)].value))
    # return data
    return worksheet_data



def read_excel_content(file_path_plus_name):
    """
    reads the content from an excel sheet that is relevant for a family tree.
    the file should consist of a first sheet that contains the names etc. and
    a second sheet that contains information about marriages / relationships
    @param file_path_plus_name: Name of file and the path to it.
    @return: List of two lists. the first list contains information about the 
    people, the second list contains information about marriages/relationships
    """
    global letters
    
    # initialize return list
    family_tree_data = [[],[]]
    # check if file exists, return empty array otherwise
    if os.path.exists(file_path_plus_name):
        # open the Excel file
        wb = load_workbook(filename=file_path_plus_name)
        # get the two sheets with relevant data
        sheet_people = wb[wb.sheetnames[0]]
        sheet_relationships = wb[wb.sheetnames[1]]
        # process worksheets
        family_tree_data[0] = extract_values_from_sheet(sheet_people)
        family_tree_data[1] = extract_values_from_sheet(sheet_relationships)
    # return data
    return family_tree_data
